"""Admin management routes"""
import base64
import io
import os
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import bcrypt

from database import get_db
from models import User, CheckIn, Location, CheckInSession
from utils.time_utils import beijing_now_naive
from schemas import (
    UserCreate, UserResponse, LocationCreate, LocationResponse,
    CheckInRecord, StatisticsResponse, LocationValidateRequest,
    AdminResetPasswordRequest, CorrectionRequest,
    SessionCreate, SessionResponse, ActiveSessionResponse, SessionUpdate,
)
from services.face_service import (
    extract_embedding, embedding_to_bytes,
    extract_embedding_from_base64, update_embedding,
)
from services.location_service import is_within_range, reverse_geocode, haversine_distance
from utils.security import decode_access_token

router = APIRouter(prefix="/api/admin", tags=["admin"])


async def get_current_admin(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth[7:]
    payload = decode_access_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    if payload.get("role") not in ("admin", "instructor"):
        raise HTTPException(status_code=403, detail="Admin or instructor required")
    return payload


# ── Users ────────────────────────────────────────────────
@router.get("/users", response_model=list[UserResponse])
async def list_users(
    role: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    stmt = select(User).where(User.is_active == True)
    if role:
        stmt = stmt.where(User.role == role)
    stmt = stmt.order_by(User.created_at.desc())
    result = await db.execute(stmt)
    users = result.scalars().all()

    return [
        UserResponse(
            id=u.id,
            username=u.username,
            name=u.name,
            role=u.role,
            has_face=u.face_embedding is not None,
            is_active=u.is_active,
            created_at=u.created_at,
        )
        for u in users
    ]


@router.post("/users", response_model=UserResponse)
async def create_user(
    req: UserCreate,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    import bcrypt

    stmt = select(User).where(User.username == req.username)
    result = await db.execute(stmt)
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Username already exists")

    user = User(
        username=req.username,
        password_hash=bcrypt.hashpw(req.password.encode(), bcrypt.gensalt()).decode(),
        name=req.name,
        role=req.role,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)

    return UserResponse(
        id=user.id,
        username=user.username,
        name=user.name,
        role=user.role,
        has_face=False,
        is_active=user.is_active,
        created_at=user.created_at,
    )


class FaceReg(BaseModel):
    face_image_base64: str


@router.post("/users/{user_id}/face")
async def register_face(
    user_id: int,
    body: FaceReg,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Register a face for a user. Body: {"face_image_base64": "..."}"""
    stmt = select(User).where(User.id == user_id)
    result = await db.execute(stmt)
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    embedding = extract_embedding_from_base64(body.face_image_base64)
    if embedding is None:
        raise HTTPException(status_code=400, detail="No face detected in image")

    user.face_embedding = embedding_to_bytes(embedding)

    # Save photo
    from services.face_service import save_checkin_photo
    photo_path = save_checkin_photo(body.face_image_base64, f"face_user_{user_id}.jpg")
    user.face_photo_path = photo_path

    await db.commit()

    return {"success": True, "message": f"Face registered for {user.name}"}


@router.post("/users/{user_id}/deactivate")
async def deactivate_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    stmt = select(User).where(User.id == user_id)
    result = await db.execute(stmt)
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = False
    await db.commit()
    return {"success": True}


# ── Locations ────────────────────────────────────────────
@router.get("/locations", response_model=list[LocationResponse])
async def list_locations(
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    stmt = select(Location).where(Location.is_active == True)
    result = await db.execute(stmt)
    locations = result.scalars().all()
    return [
        LocationResponse(
            id=l.id, name=l.name, latitude=l.latitude,
            longitude=l.longitude, radius_meters=l.radius_meters,
            is_active=l.is_active,
        )
        for l in locations
    ]


@router.post("/locations", response_model=LocationResponse)
async def create_location(
    req: LocationCreate,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    # Check for duplicate name
    dup_stmt = select(Location).where(Location.name == req.name.strip(), Location.is_active == True)
    dup_result = await db.execute(dup_stmt)
    if dup_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="签到点名称已存在")

    location = Location(
        name=req.name,
        latitude=req.latitude,
        longitude=req.longitude,
        radius_meters=req.radius_meters,
        created_by=int(_admin["sub"]),
    )
    db.add(location)
    await db.commit()
    await db.refresh(location)
    return LocationResponse(
        id=location.id, name=location.name,
        latitude=location.latitude, longitude=location.longitude,
        radius_meters=location.radius_meters, is_active=location.is_active,
    )


@router.delete("/locations/{location_id}")
async def delete_location(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    stmt = select(Location).where(Location.id == location_id)
    result = await db.execute(stmt)
    location = result.scalar_one_or_none()
    if not location:
        raise HTTPException(status_code=404, detail="签到点不存在")
    location.is_active = False
    await db.commit()
    return {"ok": True}


# ── Statistics ───────────────────────────────────────────
@router.get("/statistics", response_model=StatisticsResponse)
async def get_statistics(
    date: Optional[str] = None,
    location_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Get daily attendance statistics. date format: YYYY-MM-DD (default: today)"""
    if date:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    else:
        target_date = beijing_now_naive().date()

    day_start = datetime(target_date.year, target_date.month, target_date.day)
    day_end = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)

    # Total users
    total_stmt = select(func.count(User.id)).where(User.is_active == True, User.role == "student")
    total_result = await db.execute(total_stmt)
    total_users = total_result.scalar() or 0

    # Today's check-ins
    ci_stmt = select(CheckIn).options(selectinload(CheckIn.user)).where(
        CheckIn.check_in_time >= day_start,
        CheckIn.check_in_time <= day_end,
    )
    if location_id:
        ci_stmt = ci_stmt.where(CheckIn.location_id == location_id)
    ci_result = await db.execute(ci_stmt.order_by(CheckIn.check_in_time.desc()))
    all_checkins = ci_result.scalars().all()

    checked_in_user_ids = set(c.user_id for c in all_checkins)
    checked_in_today = len(checked_in_user_ids)
    not_checked_in = max(0, total_users - checked_in_today)

    # Average duration for completed ones
    completed = [c for c in all_checkins if c.check_out_time is not None]
    if completed:
        durations = [(c.check_out_time - c.check_in_time).total_seconds() / 60 for c in completed]
        avg_duration = sum(durations) / len(durations)
    else:
        avg_duration = None

    records = []
    for c in all_checkins[:50]:  # limit to 50
        records.append(CheckInRecord(
            id=c.id, user_id=c.user_id,
            user_name=c.user.name if c.user else f"User#{c.user_id}",
            role=c.user.role if c.user else "student",
            check_in_time=c.check_in_time,
            check_out_time=c.check_out_time,
            location_name=c.location_name,
            status=c.status,
            is_auto_checkout=c.is_auto_checkout,
            check_in_photo=f"/static/{c.check_in_photo}" if c.check_in_photo else None,
            check_out_photo=f"/static/{c.check_out_photo}" if c.check_out_photo else None,
            original_user_id=c.original_user_id,
            corrected_by=c.corrected_by,
            corrected_at=c.corrected_at,
        ))

    return StatisticsResponse(
        total_users=total_users,
        checked_in_today=checked_in_today,
        not_checked_in=not_checked_in,
        total_checkins_today=len(all_checkins),
        avg_duration_minutes=round(avg_duration, 1) if avg_duration else None,
        records=records,
    )


async def _get_current_user(request: Request):
    """Auth that accepts any valid logged-in user."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_access_token(auth[7:])
    if payload is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    return payload


@router.get("/checkins", response_model=list[CheckInRecord])
async def list_checkins(
    date: Optional[str] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _user=Depends(_get_current_user),
):
    # Students can only see their own records
    if _user.get("role") == "student":
        user_id = int(_user["sub"])
    stmt = select(CheckIn).options(selectinload(CheckIn.user))
    if date:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
        day_start = datetime(target_date.year, target_date.month, target_date.day)
        day_end = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)
        stmt = stmt.where(CheckIn.check_in_time >= day_start, CheckIn.check_in_time <= day_end)
    if user_id:
        stmt = stmt.where(CheckIn.user_id == user_id)
    if status:
        stmt = stmt.where(CheckIn.status == status)

    stmt = stmt.order_by(CheckIn.check_in_time.desc()).limit(100)
    result = await db.execute(stmt)
    checkins = result.scalars().all()

    return [
        CheckInRecord(
            id=c.id, user_id=c.user_id,
            user_name=c.user.name if c.user else f"User#{c.user_id}",
            role=c.user.role if c.user else "student",
            check_in_time=c.check_in_time,
            check_out_time=c.check_out_time,
            location_name=c.location_name,
            status=c.status,
            is_auto_checkout=c.is_auto_checkout,
            check_in_photo=f"/static/{c.check_in_photo}" if c.check_in_photo else None,
            check_out_photo=f"/static/{c.check_out_photo}" if c.check_out_photo else None,
            original_user_id=c.original_user_id,
            corrected_by=c.corrected_by,
            corrected_at=c.corrected_at,
        )
        for c in checkins
    ]


# ── Correction ────────────────────────────────────────────
@router.post("/checkins/{checkin_id}/correct")
async def correct_checkin(
    checkin_id: int,
    req: CorrectionRequest,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Admin corrects a misidentified checkin: reassign user + self-learning."""
    # 1. Load checkin record
    stmt = select(CheckIn).options(selectinload(CheckIn.user)).where(CheckIn.id == checkin_id)
    result = await db.execute(stmt)
    checkin = result.scalar_one_or_none()
    if checkin is None:
        raise HTTPException(status_code=404, detail="Checkin record not found")

    # 2. Load target (correct) user
    user_stmt = select(User).where(User.id == req.correct_user_id, User.is_active == True)
    user_result = await db.execute(user_stmt)
    correct_user = user_result.scalar_one_or_none()
    if correct_user is None:
        raise HTTPException(status_code=404, detail="Target user not found")

    # 3. Save original user if first correction
    if checkin.original_user_id is None:
        checkin.original_user_id = checkin.user_id

    # 4. Reassign
    admin_id = int(_admin["sub"])
    checkin.user_id = req.correct_user_id
    checkin.corrected_by = admin_id
    checkin.corrected_at = beijing_now_naive()

    # 5. Self-learning: extract embedding from stored photo, blend into correct user
    if checkin.check_in_photo and correct_user.face_embedding is not None:
        from config import PHOTO_DIR
        photo_path = os.path.join(PHOTO_DIR, os.path.basename(checkin.check_in_photo))
        if os.path.exists(photo_path):
            with open(photo_path, "rb") as f:
                photo_bytes = f.read()
            new_emb = extract_embedding(photo_bytes)
            if new_emb is not None:
                correct_user.face_embedding = update_embedding(correct_user.face_embedding, new_emb)

    await db.commit()
    await db.refresh(checkin)

    return {
        "success": True,
        "message": f"已将签到记录修正为 {correct_user.name}",
        "record": {
            "id": checkin.id,
            "user_id": checkin.user_id,
            "user_name": correct_user.name,
            "original_user_id": checkin.original_user_id,
            "corrected_by": checkin.corrected_by,
            "corrected_at": checkin.corrected_at.isoformat() if checkin.corrected_at else None,
        },
    }


# ── Location validation ──────────────────────────────────
@router.post("/validate-location")
async def validate_location(
    req: LocationValidateRequest,
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Location).where(Location.id == req.location_id, Location.is_active == True)
    result = await db.execute(stmt)
    location = result.scalar_one_or_none()
    if location is None:
        raise HTTPException(status_code=404, detail="Location not found")

    within, distance = is_within_range(req.lat, req.lng, location.latitude, location.longitude, location.radius_meters, req.source)
    addr = await reverse_geocode(req.lat, req.lng)

    return {
        "within_range": within,
        "distance_meters": round(distance, 1),
        "max_meters": location.radius_meters,
        "location_name": location.name,
        "address": addr,
    }


# ── Admin password reset ──────────────────────────────────
@router.post("/users/{user_id}/reset-password")
async def admin_reset_password(
    user_id: int,
    req: AdminResetPasswordRequest,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Admin resets a user's password."""
    stmt = select(User).where(User.id == user_id)
    result = await db.execute(stmt)
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    user.password_hash = bcrypt.hashpw(req.new_password.encode(), bcrypt.gensalt()).decode()
    await db.commit()

    return {"success": True, "message": f"密码已重置为: {req.new_password}"}


# ── Excel export ──────────────────────────────────────────
@router.get("/export")
async def export_checkins(
    date: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Export checkin records as Excel with two sheets: 已签到 / 未签到."""
    # ── Determine date range ──
    if date_from and date_to:
        d_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        date_label = f"{date_from} 至 {date_to}"
        day_start = datetime(d_from.year, d_from.month, d_from.day)
        day_end = datetime(d_to.year, d_to.month, d_to.day, 23, 59, 59)
    elif date:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
        date_label = target_date.strftime("%Y-%m-%d")
        day_start = datetime(target_date.year, target_date.month, target_date.day)
        day_end = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)
    else:
        target_date = beijing_now_naive().date()
        date_label = target_date.strftime("%Y-%m-%d")
        day_start = datetime(target_date.year, target_date.month, target_date.day)
        day_end = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59)

    # ── Fetch checkin records ──
    stmt = select(CheckIn).options(selectinload(CheckIn.user)).where(
        CheckIn.check_in_time >= day_start,
        CheckIn.check_in_time <= day_end,
    )
    if user_id:
        stmt = stmt.where(CheckIn.user_id == user_id)
    if status:
        stmt = stmt.where(CheckIn.status == status)
    result = await db.execute(stmt)
    records = list(result.scalars().all())

    # ── Sort: date ascending → same day by name ascending ──
    records.sort(key=lambda r: (r.check_in_time.date() if r.check_in_time else date.today(),
                                 r.user.name if r.user else ""))

    # ── Fetch all active students ──
    users_stmt = select(User).where(User.is_active == True, User.role == "student").order_by(User.name)
    users_result = await db.execute(users_stmt)
    all_students = list(users_result.scalars().all())
    student_map = {u.id: u for u in all_students}

    # ── Fetch sessions that overlap with the date range ──
    d_from_date = day_start.date()
    d_to_date = day_end.date()
    sess_stmt = (select(CheckInSession)
                 .options(selectinload(CheckInSession.location))
                 .where(CheckInSession.start_date <= d_to_date)
                 .where((CheckInSession.end_date >= d_from_date) | (CheckInSession.end_date == None))
                 .order_by(CheckInSession.created_at.desc()))
    sess_result = await db.execute(sess_stmt)
    all_sessions = list(sess_result.scalars().all())

    def _session_valid_for_date(session, check_date):
        """Check if a session is valid for a specific calendar date."""
        if session.status != "active" and session.status != "ended":
            return False
        if session.start_date and check_date < session.start_date:
            return False
        if session.end_date and check_date > session.end_date:
            return False
        if session.recurring_days:
            weekday = str(check_date.weekday())
            valid_days = [d.strip() for d in session.recurring_days.split(",") if d.strip()]
            if weekday not in valid_days:
                return False
        return True

    # ── Build per-day attendance ──
    all_dates = []
    cursor = d_from_date
    while cursor <= d_to_date:
        all_dates.append(cursor)
        cursor += timedelta(days=1)

    # Group records by date
    records_by_date = {}
    for r in records:
        rd = r.check_in_time.date() if r.check_in_time else d_from_date
        records_by_date.setdefault(rd, []).append(r)

    # Per-day absent analysis
    day_absent_data = []  # list of (date, absent_students, expected_count, session_names)
    total_absent_days = 0

    for day in all_dates:
        day_sessions = [s for s in all_sessions if _session_valid_for_date(s, day)]
        if not day_sessions:
            continue  # no sessions active this day → no expected attendance

        # Union of targeted users across all active sessions
        expected_ids = set()
        session_names = []
        for s in day_sessions:
            loc_name = s.name or (s.location.name if s.location else f"Loc#{s.location_id}")
            session_names.append(loc_name)
            if s.target_user_ids:
                for uid in s.target_user_ids.split(","):
                    uid = uid.strip()
                    if uid.isdigit():
                        expected_ids.add(int(uid))
            else:
                # null = all students
                for u in all_students:
                    expected_ids.add(u.id)

        if not expected_ids:
            continue

        day_records = records_by_date.get(day, [])
        day_checked_ids = set(r.user_id for r in day_records)
        day_absent_ids = expected_ids - day_checked_ids

        absent_students = [student_map[uid] for uid in sorted(day_absent_ids) if uid in student_map]
        if absent_students:
            total_absent_days += len(absent_students)
        day_absent_data.append({
            "date": day,
            "absent": absent_students,
            "expected": len(expected_ids),
            "checked": len(day_checked_ids & expected_ids),
            "sessions": ", ".join(session_names),
        })

    # ── Determine overall checked-in user IDs ──
    checked_in_user_ids = set()
    for r in records:
        checked_in_user_ids.add(r.user_id)

    # ── Styles ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    absent_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    day_header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Per-user color palette
    user_colors = [
        "E8F5E9", "FFF3E0", "E3F2FD", "FCE4EC", "F3E5F5",
        "E0F7FA", "FFF8E1", "EDE7F6", "EFEBE9", "E8EAF6",
        "F1F8E9", "FBE9E7", "E0F2F1", "FFF9C4", "F5F5F5",
    ]
    user_color_map = {}
    _color_idx = 0
    for r in records:
        uid = r.user_id
        if uid not in user_color_map:
            user_color_map[uid] = user_colors[_color_idx % len(user_colors)]
            _color_idx += 1

    export_time = beijing_now_naive().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ═══════════════ Sheet 1: 已签到 ═══════════════
    ws1 = wb.active
    ws1.title = "已签到"

    ws1.merge_cells("A1:K1")
    ws1["A1"] = f"已签到记录 — {date_label}"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:K2")
    ws1["A2"] = f"导出时间: {export_time}  |  签到人数: {len(checked_in_user_ids)}  |  总记录: {len(records)}"
    ws1["A2"].alignment = Alignment(horizontal="center")

    checked_in_headers = ["序号", "姓名", "用户名", "角色", "签到时间", "签退时间", "签到点", "时长(分钟)", "状态", "签到方式"]
    for col, h in enumerate(checked_in_headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(records, 1):
        user = r.user
        duration = ""
        if r.check_out_time:
            duration = round((r.check_out_time - r.check_in_time).total_seconds() / 60, 1)
        method = "自动签退" if r.is_auto_checkout else ("进行中" if r.status == "active" else "手动签退")
        status_text = "进行中" if r.status == "active" else "已完成"

        row_data = [
            i,
            user.name if user else f"User#{r.user_id}",
            user.username if user else "",
            user.role if user else "student",
            r.check_in_time.strftime("%Y-%m-%d %H:%M:%S") if r.check_in_time else "",
            r.check_out_time.strftime("%Y-%m-%d %H:%M:%S") if r.check_out_time else "",
            r.location_name or "",
            duration,
            status_text,
            method,
        ]
        row_fill = PatternFill(start_color=user_color_map.get(r.user_id, "FFFFFF"),
                               end_color=user_color_map.get(r.user_id, "FFFFFF"), fill_type="solid")
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i + 4, column=col, value=val)
            cell.border = thin_border
            cell.fill = row_fill

    col_widths_1 = [6, 10, 12, 8, 20, 20, 18, 12, 8, 12]
    for col, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ═══════════════ Sheet 2: 未签到 (per-day) ═══════════════
    ws2 = wb.create_sheet("未签到")

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"未签到明细 — {date_label}"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.merge_cells("A2:F2")
    total_absent_unique = len(set(
        uid for d in day_absent_data for u in d["absent"] for uid in [u.id]
    )) if day_absent_data else 0
    ws2["A2"] = f"导出时间: {export_time}  |  涉及天数: {len(day_absent_data)}  |  缺勤人次: {total_absent_days}"
    ws2["A2"].alignment = Alignment(horizontal="center")

    absent_headers = ["日期", "序号", "姓名", "用户名", "角色", "关联任务"]
    for col, h in enumerate(absent_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 5
    for day_data in day_absent_data:
        day_str = day_data["date"].strftime("%Y-%m-%d")
        sessions_str = day_data["sessions"]

        if not day_data["absent"]:
            # Show day with no absences
            cell = ws2.cell(row=row, column=1, value=day_str)
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=2, value="—")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=3, value="全员到齐")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=4, value="")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=5, value="")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=6, value=sessions_str)
            cell.border = thin_border; cell.fill = day_header_fill
            row += 1
        else:
            for idx, u in enumerate(day_data["absent"], 1):
                row_data = [day_str if idx == 1 else "", idx, u.name, u.username, u.role, sessions_str if idx == 1 else ""]
                for col, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    if idx % 2 == 0:
                        cell.fill = absent_fill
                row += 1

    if not day_absent_data:
        ws2.cell(row=5, column=1, value="所选日期范围内无签到任务").border = thin_border

    col_widths_2 = [14, 6, 12, 14, 8, 24]
    for col, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Stream response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"签到记录_{date_label}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ── Check-In Sessions ─────────────────────────────────────
from datetime import date, time

WEEKDAY_MAP = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6}

def _parse_date(s: Optional[str]) -> Optional[date]:
    if not s: return None
    return date.fromisoformat(s)

def _parse_time(s: Optional[str]) -> Optional[time]:
    if not s: return None
    return time.fromisoformat(s)

def is_session_time_valid(session: CheckInSession) -> bool:
    """Check if current Beijing time falls within the session's time window."""
    now = beijing_now_naive()
    today = now.date()
    now_t = now.time()

    if session.start_date and today < session.start_date:
        return False
    if session.end_date and today > session.end_date:
        return False

    if session.recurring_days:
        weekday = str(today.weekday())  # 0=Mon ... 6=Sun
        valid_days = [d.strip() for d in session.recurring_days.split(",") if d.strip()]
        if weekday not in valid_days:
            return False

    if session.checkin_start_time and now_t < session.checkin_start_time:
        return False
    if session.checkin_end_time and now_t > session.checkin_end_time:
        return False

    return True


def _session_to_response(session: CheckInSession, target_user_names: Optional[str] = None) -> dict:
    return {
        "id": session.id,
        "location_id": session.location_id,
        "location_name": session.location.name if session.location else None,
        "name": session.name,
        "created_by": session.created_by,
        "creator_name": session.creator.name if session.creator else None,
        "status": session.status,
        "created_at": session.created_at,
        "ended_at": session.ended_at,
        "start_date": session.start_date.isoformat() if session.start_date else None,
        "end_date": session.end_date.isoformat() if session.end_date else None,
        "checkin_start_time": session.checkin_start_time.strftime("%H:%M") if session.checkin_start_time else None,
        "checkin_end_time": session.checkin_end_time.strftime("%H:%M") if session.checkin_end_time else None,
        "recurring_days": session.recurring_days,
        "target_user_ids": session.target_user_ids,
        "target_user_names": target_user_names,
        "time_valid": is_session_time_valid(session),
    }


async def _resolve_target_names(db: AsyncSession, target_user_ids: Optional[str]) -> Optional[str]:
    """Resolve comma-separated user IDs to comma-separated user names."""
    if not target_user_ids:
        return None
    ids = [int(x.strip()) for x in target_user_ids.split(",") if x.strip().isdigit()]
    if not ids:
        return None
    stmt = select(User.name).where(User.id.in_(ids))
    result = await db.execute(stmt)
    names = [row[0] for row in result.fetchall()]
    return ", ".join(names)


@router.post("/sessions", response_model=SessionResponse)
async def create_session(
    req: SessionCreate,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Start a new check-in session. Multiple sessions can coexist."""
    loc_stmt = select(Location).where(Location.id == req.location_id, Location.is_active == True)
    loc_result = await db.execute(loc_stmt)
    location = loc_result.scalar_one_or_none()
    if not location:
        raise HTTPException(status_code=404, detail="签到点不存在或已停用")

    # Check for duplicate name among active sessions
    if req.name and req.name.strip():
        dup_stmt = select(CheckInSession).where(
            CheckInSession.name == req.name.strip(),
            CheckInSession.status == "active",
        )
        dup_result = await db.execute(dup_stmt)
        if dup_result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="签到任务名称已存在，请使用不同的名称")

    session = CheckInSession(
        location_id=req.location_id,
        created_by=int(_admin["sub"]),
        status="active",
        name=req.name.strip() if req.name else None,
        start_date=_parse_date(req.start_date),
        end_date=_parse_date(req.end_date),
        checkin_start_time=_parse_time(req.checkin_start_time),
        checkin_end_time=_parse_time(req.checkin_end_time),
        recurring_days=req.recurring_days.strip() if req.recurring_days else None,
        target_user_ids=req.target_user_ids.strip() if req.target_user_ids else None,
    )
    db.add(session)
    await db.commit()
    stmt = (select(CheckInSession).where(CheckInSession.id == session.id)
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator)))
    result = await db.execute(stmt)
    session = result.scalar_one()
    names = await _resolve_target_names(db, session.target_user_ids)
    return _session_to_response(session, target_user_names=names)


@router.post("/sessions/{session_id}/end", response_model=SessionResponse)
async def end_session(
    session_id: int,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """End an active check-in session."""
    stmt = (select(CheckInSession).where(CheckInSession.id == session_id)
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator)))
    result = await db.execute(stmt)
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="签到任务不存在")
    if session.status != "active":
        raise HTTPException(status_code=400, detail="该签到任务已结束")

    session.status = "ended"
    session.ended_at = beijing_now_naive()
    await db.commit()
    return _session_to_response(session)


@router.put("/sessions/{session_id}", response_model=SessionResponse)
async def update_session(
    session_id: int,
    req: SessionUpdate,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Update an existing check-in session. location_id cannot be changed."""
    stmt = (select(CheckInSession).where(CheckInSession.id == session_id)
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator)))
    result = await db.execute(stmt)
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="签到任务不存在")

    # Check for duplicate name (exclude self)
    if req.name is not None and req.name.strip():
        dup_stmt = select(CheckInSession).where(
            CheckInSession.name == req.name.strip(),
            CheckInSession.status == "active",
            CheckInSession.id != session_id,
        )
        dup_result = await db.execute(dup_stmt)
        if dup_result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="签到任务名称已存在，请使用不同的名称")

    # 只更新提供的字段，不修改 location_id
    if req.name is not None:
        session.name = req.name.strip() if req.name else None
    if req.start_date is not None:
        session.start_date = _parse_date(req.start_date)
    if req.end_date is not None:
        session.end_date = _parse_date(req.end_date)
    if req.checkin_start_time is not None:
        session.checkin_start_time = _parse_time(req.checkin_start_time)
    if req.checkin_end_time is not None:
        session.checkin_end_time = _parse_time(req.checkin_end_time)
    if req.recurring_days is not None:
        session.recurring_days = req.recurring_days.strip() if req.recurring_days else None
    if req.target_user_ids is not None:
        session.target_user_ids = req.target_user_ids.strip() if req.target_user_ids else None

    await db.commit()
    return _session_to_response(session)


@router.get("/sessions/{session_id}/export")
async def export_session_checkins(
    session_id: int,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Export checkin records for a single session as Excel with two sheets: 已签到 / 未签到."""
    # ── 查找任务 ──
    stmt = (select(CheckInSession).where(CheckInSession.id == session_id)
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator)))
    result = await db.execute(stmt)
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="签到任务不存在")

    session_name = session.name or f"Session#{session_id}"

    # ── 确定日期范围 ──
    today = beijing_now_naive().date()
    if session.start_date:
        d_from = session.start_date
    else:
        d_from = today
    if session.end_date:
        d_to = min(session.end_date, today)
    else:
        d_to = today

    date_label = f"{d_from.isoformat()} 至 {d_to.isoformat()}"
    day_start = datetime(d_from.year, d_from.month, d_from.day)
    day_end = datetime(d_to.year, d_to.month, d_to.day, 23, 59, 59)

    # ── 解析目标任务用户 ──
    target_ids = None
    if session.target_user_ids:
        target_ids = set()
        for uid in session.target_user_ids.split(","):
            uid = uid.strip()
            if uid.isdigit():
                target_ids.add(int(uid))

    # ── 获取签到记录 ──
    ci_stmt = select(CheckIn).options(selectinload(CheckIn.user)).where(
        CheckIn.check_in_time >= day_start,
        CheckIn.check_in_time <= day_end,
    )
    ci_result = await db.execute(ci_stmt)
    records = list(ci_result.scalars().all())

    # ── 排序：日期升序 → 同日按姓名升序 ──
    records.sort(key=lambda r: (r.check_in_time.date() if r.check_in_time else d_from,
                                 r.user.name if r.user else ""))

    # ── 获取所有活跃学生 ──
    users_stmt = select(User).where(User.is_active == True, User.role == "student").order_by(User.name)
    users_result = await db.execute(users_stmt)
    all_students = list(users_result.scalars().all())
    student_map = {u.id: u for u in all_students}

    # ── 按天构建出勤数据（仅此任务） ──
    all_dates = []
    cursor = d_from
    while cursor <= d_to:
        all_dates.append(cursor)
        cursor += timedelta(days=1)

    # 按日期分组记录
    records_by_date = {}
    for r in records:
        rd = r.check_in_time.date() if r.check_in_time else d_from
        records_by_date.setdefault(rd, []).append(r)

    def _session_valid_for_date(sess, check_date):
        """检查任务在某天是否有效。"""
        if sess.status != "active" and sess.status != "ended":
            return False
        if sess.start_date and check_date < sess.start_date:
            return False
        if sess.end_date and check_date > sess.end_date:
            return False
        if sess.recurring_days:
            weekday = str(check_date.weekday())
            valid_days = [d.strip() for d in sess.recurring_days.split(",") if d.strip()]
            if weekday not in valid_days:
                return False
        return True

    # 每日缺勤分析
    day_absent_data = []
    total_absent_days = 0

    for day in all_dates:
        if not _session_valid_for_date(session, day):
            continue  # 该天任务不生效

        # 本任务的目标用户
        if target_ids is not None:
            expected_ids = target_ids.copy()
        else:
            expected_ids = set(u.id for u in all_students)

        if not expected_ids:
            continue

        day_records = records_by_date.get(day, [])
        day_checked_ids = set(r.user_id for r in day_records)
        day_absent_ids = expected_ids - day_checked_ids

        absent_students = [student_map[uid] for uid in sorted(day_absent_ids) if uid in student_map]
        if absent_students:
            total_absent_days += len(absent_students)
        day_absent_data.append({
            "date": day,
            "absent": absent_students,
            "expected": len(expected_ids),
            "checked": len(day_checked_ids & expected_ids),
            "sessions": session_name,  # 仅此任务
        })

    # ── 整体已签到用户 ──
    checked_in_user_ids = set()
    for r in records:
        checked_in_user_ids.add(r.user_id)

    # ── 样式 ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    absent_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    day_header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # 每人颜色
    user_colors = [
        "E8F5E9", "FFF3E0", "E3F2FD", "FCE4EC", "F3E5F5",
        "E0F7FA", "FFF8E1", "EDE7F6", "EFEBE9", "E8EAF6",
        "F1F8E9", "FBE9E7", "E0F2F1", "FFF9C4", "F5F5F5",
    ]
    user_color_map = {}
    _color_idx = 0
    for r in records:
        uid = r.user_id
        if uid not in user_color_map:
            user_color_map[uid] = user_colors[_color_idx % len(user_colors)]
            _color_idx += 1

    export_time = beijing_now_naive().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()

    # ═══════════════ Sheet 1: 已签到 ═══════════════
    ws1 = wb.active
    ws1.title = "已签到"

    ws1.merge_cells("A1:K1")
    ws1["A1"] = f"已签到记录 — {session_name} ({date_label})"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:K2")
    ws1["A2"] = f"导出时间: {export_time}  |  签到人数: {len(checked_in_user_ids)}  |  总记录: {len(records)}"
    ws1["A2"].alignment = Alignment(horizontal="center")

    checked_in_headers = ["序号", "姓名", "用户名", "角色", "签到时间", "签退时间", "签到点", "时长(分钟)", "状态", "签到方式"]
    for col, h in enumerate(checked_in_headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(records, 1):
        user = r.user
        duration = ""
        if r.check_out_time:
            duration = round((r.check_out_time - r.check_in_time).total_seconds() / 60, 1)
        method = "自动签退" if r.is_auto_checkout else ("进行中" if r.status == "active" else "手动签退")
        status_text = "进行中" if r.status == "active" else "已完成"

        row_data = [
            i,
            user.name if user else f"User#{r.user_id}",
            user.username if user else "",
            user.role if user else "student",
            r.check_in_time.strftime("%Y-%m-%d %H:%M:%S") if r.check_in_time else "",
            r.check_out_time.strftime("%Y-%m-%d %H:%M:%S") if r.check_out_time else "",
            r.location_name or "",
            duration,
            status_text,
            method,
        ]
        row_fill = PatternFill(start_color=user_color_map.get(r.user_id, "FFFFFF"),
                               end_color=user_color_map.get(r.user_id, "FFFFFF"), fill_type="solid")
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i + 4, column=col, value=val)
            cell.border = thin_border
            cell.fill = row_fill

    col_widths_1 = [6, 10, 12, 8, 20, 20, 18, 12, 8, 12]
    for col, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ═══════════════ Sheet 2: 未签到 (按天) ═══════════════
    ws2 = wb.create_sheet("未签到")

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"未签到明细 — {session_name} ({date_label})"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.merge_cells("A2:F2")
    total_absent_unique = len(set(
        uid for d in day_absent_data for u in d["absent"] for uid in [u.id]
    )) if day_absent_data else 0
    ws2["A2"] = f"导出时间: {export_time}  |  涉及天数: {len(day_absent_data)}  |  缺勤人次: {total_absent_days}"
    ws2["A2"].alignment = Alignment(horizontal="center")

    absent_headers = ["日期", "序号", "姓名", "用户名", "角色", "关联任务"]
    for col, h in enumerate(absent_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 5
    for day_data in day_absent_data:
        day_str = day_data["date"].strftime("%Y-%m-%d")
        sessions_str = day_data["sessions"]

        if not day_data["absent"]:
            cell = ws2.cell(row=row, column=1, value=day_str)
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=2, value="—")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=3, value="全员到齐")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=4, value="")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=5, value="")
            cell.border = thin_border; cell.fill = day_header_fill
            cell = ws2.cell(row=row, column=6, value=sessions_str)
            cell.border = thin_border; cell.fill = day_header_fill
            row += 1
        else:
            for idx, u in enumerate(day_data["absent"], 1):
                row_data = [day_str if idx == 1 else "", idx, u.name, u.username, u.role, sessions_str if idx == 1 else ""]
                for col, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    if idx % 2 == 0:
                        cell.fill = absent_fill
                row += 1

    if not day_absent_data:
        ws2.cell(row=5, column=1, value="所选日期范围内无签到任务").border = thin_border

    col_widths_2 = [14, 6, 12, 14, 8, 24]
    for col, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── 流式响应 ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{session_name}_签到记录_{date_label}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/sessions/active", response_model=ActiveSessionResponse)
async def get_active_session(
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Get all active check-in sessions (for admin dashboard)."""
    stmt = (select(CheckInSession).where(CheckInSession.status == "active")
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator))
            .order_by(CheckInSession.created_at.desc()))
    result = await db.execute(stmt)
    sessions = result.scalars().all()
    return {
        "has_active_session": len(sessions) > 0,
        "sessions": [_session_to_response(s) for s in sessions],
    }


@router.get("/sessions/{session_id}")
async def get_session(
    session_id: int,
    db: AsyncSession = Depends(get_db),
    _admin=Depends(get_current_admin),
):
    """Get a single check-in session with details."""
    stmt = (select(CheckInSession).where(CheckInSession.id == session_id)
            .options(selectinload(CheckInSession.location), selectinload(CheckInSession.creator)))
    result = await db.execute(stmt)
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="签到任务不存在")
    names = await _resolve_target_names(db, session.target_user_ids)
    return _session_to_response(session, target_user_names=names)
